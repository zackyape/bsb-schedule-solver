"""
excel_io.py

Excel reader/writer for BSB Schedule Solver v2
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from config import (
    INPUT_FILE,
    OUTPUT_FILE,
    LOCK_VALUES,
    EMPTY_VALUES,
    IGNORE_PERSONNEL,
    excel,
)
from models import (
    Assignment,
    DaySchedule,
    Person,
    WorkbookData,
    SolverResult,
)


class ExcelIO:
    def __init__(self, input_file: Path = INPUT_FILE):
        self.input_file = Path(input_file)
        self.workbook = load_workbook(self.input_file)
        self.sheet = (
            self.workbook.active
            if excel.auto_detect_sheet
            else self.workbook[excel.sheet_name]
        )

    def load(self) -> WorkbookData:
        people = self._read_people()
        days = self._read_days(people)
        return WorkbookData(people=people, days=days)

    def save(
        self,
        result: SolverResult,
        output_file: Path = OUTPUT_FILE,
    ):
        for day, assigns in result.assignments.items():
            col = excel.first_day_column + day - 1
            for person, label in assigns.items():
                row = result.row_lookup[person]
                self.sheet.cell(row=row, column=col).value = label
        self.workbook.save(output_file)

    def _read_people(self):
        people = {}
        header = self._find_header()
        row = header + 1
        while True:
            value = self.sheet.cell(row=row, column=excel.personnel_column).value
            if value is None:
                break
            name = str(value).strip()
            if name and name not in IGNORE_PERSONNEL:
                people[name] = Person(name=name, row=row)
            row += 1
        return people

    def _read_days(self, people):
        days = []
        for col in range(excel.first_day_column, excel.last_day_column + 1):
            day = self.sheet.cell(row=4, column=col).value
            if not isinstance(day, int):
                continue
            ds = DaySchedule(day=day)
            for person in people.values():
                value = self.sheet.cell(row=person.row, column=col).value
                if value in LOCK_VALUES:
                    ds.personnel.append(person.name)
                    ds.locked[person.name] = value
                    person.assignments.append(Assignment(day, value, True))
                elif value in EMPTY_VALUES:
                    ds.personnel.append(person.name)
                    ds.empty.append(person.name)
                    person.assignments.append(Assignment(day, None, False))
            days.append(ds)
        return days

    def _find_header(self):
        for r in range(1, self.sheet.max_row + 1):
            v = self.sheet.cell(row=r, column=1).value
            if v and str(v).strip().upper() == "NAMA / TANGGAL":
                return r
        raise RuntimeError("Header 'NAMA / TANGGAL' tidak ditemukan.")

# === excel_io.py (Part 2) ===
# Append this below Part 1

    def _parse_schedule(
        self,
        people: dict[str, Person],
        schedules: list[DaySchedule],
    ) -> None:
        """
        Parse every assignment from Excel.
        """

        day_map = {d.day: d for d in schedules}

        for column in range(
            config.excel.first_day_column,
            config.excel.last_day_column + 1,
        ):

            day = self.sheet.cell(
                row=config.excel.header_row,
                column=column,
            ).value

            if day not in day_map:
                continue

            schedule = day_map[day]

            for person in people.values():

                value = self.sheet.cell(
                    row=person.row,
                    column=column,
                ).value

                if value is None:
                    continue

                if value in config.labels.lock_values:

                    schedule.personnel.append(person.name)
                    schedule.locked[person.name] = value

                    person.add_assignment(
                        Assignment(
                            day=day,
                            label=value,
                            locked=True,
                        )
                    )

                elif value in config.labels.empty_values:

                    schedule.personnel.append(person.name)
                    schedule.empty.append(person.name)

                    person.add_assignment(
                        Assignment(
                            day=day,
                            label=None,
                            locked=False,
                        )
                    )

    def load(self) -> WorkbookData:

        people = self._find_people()

        schedules = self._find_days(people)

        self._parse_schedule(
            people,
            schedules,
        )

        row_lookup = {
            p.name: p.row
            for p in people.values()
        }

        day_lookup = {
            d.day: index
            for index, d in enumerate(schedules)
        }

        return WorkbookData(
            people=people,
            schedules=schedules,
            row_lookup=row_lookup,
            day_lookup=day_lookup,
        )

# === excel_io.py (Part 3) ===
# Append below Part 2

from openpyxl.styles import Alignment


class ExcelWriter:
    """
    Write solver result back to Excel.
    """

    def __init__(self, reader: ExcelReader):
        self.reader = reader
        self.workbook = reader.workbook
        self.sheet = reader.sheet

    def save(self, result) -> None:

        for day, assignments in result.assignments.items():

            column = (
                config.excel.first_day_column
                + day
                - 1
            )

            for person, label in assignments.items():

                row = result.row_lookup[person]

                self.sheet.cell(
                    row=row,
                    column=column,
                ).value = label

        self._format_sheet()

        self.workbook.save(
            config.files.output_excel
        )

    def _format_sheet(self):

        for column_cells in self.sheet.columns:

            length = 0

            letter = column_cells[0].column_letter

            for cell in column_cells:

                if cell.value is None:
                    continue

                length = max(
                    length,
                    len(str(cell.value)),
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

            self.sheet.column_dimensions[
                letter
            ].width = max(10, length + 2)


def load_workbook_data(
    path=None,
):
    """
    Convenience helper.
    """

    reader = ExcelReader(path)

    return reader.load()


def save_solver_result(
    reader: ExcelReader,
    result,
):
    """
    Convenience helper.
    """

    writer = ExcelWriter(reader)

    writer.save(result)
    
